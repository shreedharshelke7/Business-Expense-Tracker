import pandas as pd 
import openpyxl
import os
from  datetime import datetime
import matplotlib.pyplot as plt
from fpdf import FPDF

def clean_data(df):
    df['Date'] = pd.to_datetime(df['Date'],format='%d-%m-%Y') # converting date colunm type
    # stripping out extra space
    df['Department'] =  df['Department'].str.strip()
    df['Category'] = df['Category'].str.strip()
    df['Description'] = df['Description'].str.strip()
    
    # Formatting columns department and category
    df['Department'] = df['Department'].str.title()
    df['Category'] = df['Category'].str.title()

    # droping any row with empty amount 
    df= df.dropna()
    return df

def spend_dept_month(df):
        result_df=df.groupby(['Month','Department'])['Amount'].sum().reset_index(name="amount spent department/month")
        return result_df 

def spend_category_month(df):
    result_df=df.groupby(['Month','Category'])['Amount'].sum().reset_index(name="amount spent category/Month")
    return result_df

def calculate_variance(df):
    Actual = df.groupby(['Month','Category'])['Amount'].sum().reset_index(name = "Actual")
    Budget=df.groupby(['Month','Category'])['Budget'].mean().reset_index()
    variance_df=pd.merge(Budget,Actual, on=['Month','Category'])
    variance_df['Variance'] = variance_df['Budget'] - variance_df['Actual'] 
    return variance_df

def anomaly_detection(df):
    mean_std_df = df.groupby(['Category'])['Amount'].agg(['mean','std']).reset_index()
    df = pd.merge(df,mean_std_df ,on=['Category'])
    df['is_anamoly'] = df['Amount'] > df['mean'] + 2*df['std']              
    anomaly_df = df[df['is_anamoly']==True]
    return anomaly_df

def create_excel(variance_df,anomaly_df):
    os.makedirs('outputs',exist_ok=True)
    date_time=str(datetime.now().strftime("%d-%m-%Y"))
    filename = os.path.join( 'outputs','expense_report_'+date_time+'.xlsx')

    workbook = openpyxl.Workbook()
    sheet1 = workbook.active
    sheet1.title ="summary"
    sheet2=workbook.create_sheet("anamolies")
    sheet2.title = "anamolies"
    workbook.save(filename)
    
    col_no=1
    for column_no in range(0,len(variance_df.columns)):
        sheet1.cell(row=1,column=col_no,value=variance_df.columns[column_no])
        column_name=variance_df.columns[column_no]
        column_data=variance_df[column_name]
        for row_no , entry_data in enumerate(column_data ,start=2):
            sheet1.cell(row=row_no,column=col_no,value=entry_data)
        col_no+=1
        
    col_no=1
    for column_no in range(0,len(anomaly_df.columns)):
        sheet2.cell(row=1,column=col_no,value=anomaly_df.columns[column_no])
        column_name=anomaly_df.columns[column_no]
        column_data=anomaly_df[column_name]
        for row_no , entry_data in enumerate(column_data ,start=2):
            sheet2.cell(row=row_no,column=col_no,value=entry_data)
        col_no+=1
    workbook.save(filename)
    workbook.close()

def generate_chart(variance_df,dept_month_df):

    chart1_df=dept_month_df.pivot(index='Month',columns='Department',values='amount spent department/month')
    chart1_df.plot(kind='bar',color=['#3498db', "#28d16f","#ddbe35"],edgecolor='black')
    plt.title("expense department per Month")
    plt.ylabel("expense")
    plt.xlabel("Month")
    plt.xticks(rotation=0)
    plt.grid(axis='y',linestyle='--',alpha=0.5)
    plt.savefig('outputs/chart1.png',bbox_inches='tight')
    chart2_actual=variance_df.groupby('Month')['Actual'].sum().reset_index()
    chart2_budget=variance_df.groupby('Month')['Budget'].sum().reset_index()
    fig,ax= plt.subplots(figsize=(10,10))
    ax.plot(chart2_actual['Month'],chart2_actual['Actual'],label='Actual expense',color="#d85011",marker='o')
    ax.plot(chart2_budget['Month'],chart2_budget['Budget'],label='Budget',color="#61d811",marker='s')
    ax.set_title("Actual V/s Budget")
    ax.set_xlabel("Month")
    ax.set_ylabel("Amount")
    ax.legend(loc='upper right')
    plt.savefig('outputs/chart2.png',bbox_inches='tight')

def generate_pdf_report(anomaly_df):
    pdf = FPDF()
    #==================== report cover ===============
    pdf.add_page()
    pdf.set_margins(0, 0, 0)
    pdf.set_fill_color(223,223,223)
    pdf.rect(0, 0, 210, 297, style="F")

    #====== header
    pdf.set_fill_color(8,207,238)
    pdf.set_x(30)
    pdf.set_font('helvetica',"B",size=12)
    pdf.cell(20,40," ",fill=True,new_x="LMARGIN",new_y="NEXT")
    pdf.ln(3)

    #===== LOGO
    pdf.set_fill_color(223,223,223)
    pdf.image('Higgs_logo.jpeg',x=30,w=20,h=20)
    pdf.ln(50)

    #=====Body
    pdf.set_font('helvetica',"B",size=32)
    pdf.set_text_color(8,207,238)
    pdf.set_x(30)
    pdf.multi_cell(0,10,"Annual\nFinancial Report",new_x="LMARGIN",new_y="NEXT")
    pdf.set_fill_color(8,207,238)
    pdf.ln(1)
    pdf.set_x(30)
    pdf.cell(100,1," ",fill=True,new_x="LMARGIN",new_y="NEXT")

    curr_year = str(datetime.now().strftime("%Y"))
    pdf.set_text_color(0,0,0)
    pdf.set_font('helvetica',"B",size=10)
    pdf.set_fill_color(223,223,223)
    pdf.set_x(30)
    pdf.cell(5,5,curr_year,fill=True,new_x="LMARGIN",new_y="NEXT")
    pdf.ln(20)
    pdf.set_fill_color(8,207,238)
    pdf.cell(0,100," ",fill=True)

    pdf.set_xy(5,250)
    pdf.cell(5,5,"report by-Higgsbosson",fill=False,new_x="LMARGIN",new_y="NEXT")

    pdf.set_xy(30,170)
    pdf.multi_cell(0,5,"[company Name]\n[prepared by]\n[specific Title]",new_x="LMARGIN",new_y="NEXT")

    #====================== chart 1  ======================
    pdf.add_page()
    pdf.ln(10)
    pdf.set_fill_color(243,231,181)
    pdf.set_font("helvetica","B",size=32)

    pdf.multi_cell(0,20,"\t\tExpense by Department per Month",fill=True,new_x="LMARGIN",new_y="NEXT")
    pdf.image('outputs\chart1.png',x=30,y=50,w=150,h=150)
    pdf.ln(10)


    #====================== chart 2 ======================
    pdf.add_page()
    pdf.ln(10)
    pdf.set_fill_color(243,231,181)
    pdf.set_font("helvetica","B",size=32)

    pdf.multi_cell(0,20,"\t\tBudget V/s Actual Expense",fill=True,new_x="LMARGIN",new_y="NEXT")
    pdf.image('outputs\chart2.png',x=10,y=50,w=170,h=160)
    #================== anamolies ===========================
    
    if anomaly_df is not None:
        anomaly_df['Date']= anomaly_df['Date'].dt.strftime('%d-%m-%Y')
        anomaly_df[['std','mean']]= anomaly_df[['std','mean']].round(2)
        anomaly_df.drop('Description',axis=1,inplace=True)
        pdf.add_page()
        pdf.ln(20)
        pdf.set_fill_color(243,231,181)
        pdf.set_font("helvetica","B",size=32)
        pdf.cell(0,20,"\t\tAnamolies Found",fill=True,new_x="LMARGIN",new_y="NEXT")
        pdf.ln()
        pdf.set_font("helvetica",size=10)
        pdf.set_x(20)
        for col in anomaly_df.columns:
            
            pdf.cell(20,10,col,border=1)
        pdf.ln()
        pdf.set_font("helvetica",size=8)
        for index,row in anomaly_df.iterrows():
            pdf.set_x(20)
            pdf.cell(20,10,str(row['Date']),border=1)
            pdf.cell(20,10,str(row['Department']),border=1)
            pdf.cell(20,10,str(row['Category']),border=1)
            pdf.cell(20,10,str(row['Amount']),border=1)
            pdf.cell(20,10,str(row['Budget']),border=1)
            pdf.cell(20,10,str(row['Month']),border=1)
            pdf.cell(20,10,str(row['mean']),border=1)
            pdf.cell(20,10,str(row['std']),border=1)
            pdf.cell(20,10,str(row['is_anamoly']),border=1)
            pdf.ln()
    

    #==== output =================
    pdf.output("outputs\ report.pdf")
   




def main():
    df = pd.read_csv('data/business_expense_tracker_data.csv')
    df = clean_data(df)
    df['Month'] = df['Date'].dt.month_name()
    print("✓ Data loaded and cleaned")
    dept_month_df=spend_dept_month(df)             # spending by department per month
    category_month_df=spend_category_month(df)     # spending by Category per month | Actual Spending
    variance_df =calculate_variance(df)            # variance = Budget - actual
    print("✓ Analysis complete")
    anomaly_df=anomaly_detection(df)               # gives df having STD and mean 
    print("✓ Anomalies detected")
    create_excel(variance_df,anomaly_df)           # create excel sheet report
    print("✓ Excel exported")
    generate_chart(variance_df,dept_month_df)      # generate charts 
    print("✓ Charts generated")
    generate_pdf_report(anomaly_df)                # generate PDF report 
    print("✓ PDF report created")    
main()